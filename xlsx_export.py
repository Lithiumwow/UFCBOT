"""
Builds the event recap spreadsheet for /spread-sheet -- Pick / Opponent /
Result / Odds / Unit Bet / Unit Profit / Cash Profit / ROI, grouped into
Straight Picks / Props / Parlays sections with subtotal rows and a grand
total, matching the layout the user provided as a reference.

Two real data-model limitations, both noted in the sheet itself via a
comment rather than hidden:
- Only the OVERALL bet's odds/stake are stored, not per-leg odds within a
  parlay -- so a parlay's individual leg rows show Pick/Opponent/Result
  but no odds; only the combined "Parlay Odds" row has real odds/profit
  numbers, matching what's actually in the database.
- Per-leg Won/Loss is only tracked when a leg was auto-graded (see
  grading.py). A leg on a bet you settled manually via the Won/Loss
  buttons shows blank/pending here even though the overall bet is
  settled, since manual settling only updates the bet as a whole.

Formulas, not hardcoded results, per the xlsx skill: profit/ROI cells
recompute from Odds/Unit Bet/Result if edited, rather than baking in
Python-computed numbers.
"""
from __future__ import annotations

from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from grading import _name_matches  # reused for fuzzy opponent lookup

FONT_NAME = "Arial"

HEADER_FILL = PatternFill("solid", fgColor="D9D9D9")
SECTION_FILL = PatternFill("solid", fgColor="BFBFBF")
TOTAL_FILL = PatternFill("solid", fgColor="C6EFCE")
WIN_FONT = Font(name=FONT_NAME, color="006100")
LOSS_FONT = Font(name=FONT_NAME, color="9C0006")
BOLD = Font(name=FONT_NAME, bold=True)
TITLE_FONT = Font(name=FONT_NAME, bold=True, size=14)
INPUT_FONT = Font(name=FONT_NAME, color="0000FF")

THIN = Side(style="thin", color="B7B7B7")
BORDER = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)

COLUMNS = ["Pick", "Opponent", "Result", "Odds", "Unit Bet", "Unit Profit", "Cash Profit", "ROI"]
COL_LETTERS = {name: get_column_letter(i + 1) for i, name in enumerate(COLUMNS)}


def _find_opponent(fighter: Optional[str], fights: list[tuple[str, str]]) -> str:
    if not fighter:
        return ""
    for fight in fights:
        a, b = fight[0], fight[1]
        if _name_matches(fighter, a):
            return b
        if _name_matches(fighter, b):
            return a
    return ""


def _leg_category(bet: dict[str, Any], legs: list[dict[str, Any]]) -> str:
    if len(legs) > 1:
        return "Parlay"
    if legs and (legs[0].get("outcome_type") or "ML").upper() == "ML" and legs[0].get("fighter_pick"):
        return "Straight Pick"
    return "Prop"


def _write_header_row(ws, row: int) -> None:
    for col_name in COLUMNS:
        cell = ws[f"{COL_LETTERS[col_name]}{row}"]
        cell.value = col_name
        cell.font = BOLD
        cell.fill = HEADER_FILL
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center")


def _write_pick_row(
    ws, row: int, *, pick: str, opponent: str, result: Optional[str],
    odds: Optional[int], units: Optional[float], unit_value_cell: str,
) -> None:
    ws[f"A{row}"] = pick
    ws[f"B{row}"] = opponent
    ws[f"C{row}"] = result or ""
    if result == "W":
        ws[f"C{row}"].font = WIN_FONT
    elif result == "L":
        ws[f"C{row}"].font = LOSS_FONT

    if odds is not None:
        ws[f"D{row}"] = odds
        ws[f"D{row}"].font = INPUT_FONT
    if units is not None:
        ws[f"E{row}"] = units
        ws[f"E{row}"].font = INPUT_FONT

    # Formulas -- recompute if Result/Odds/Unit Bet are edited, rather than
    # baking in a Python-computed number. Guarded against blank Odds/Unit
    # Bet (parlay leg rows intentionally leave these blank -- no per-leg
    # odds stored, see module docstring).
    ws[f"F{row}"] = (
        f'=IF(OR(D{row}="",E{row}=""),"",'
        f'IF(C{row}="W",IF(D{row}>0,E{row}*D{row}/100,E{row}*100/ABS(D{row})),'
        f'IF(C{row}="L",-E{row},"")))'
    )
    ws[f"G{row}"] = f'=IF(F{row}="","",F{row}*{unit_value_cell})'
    ws[f"H{row}"] = f'=IF(OR(E{row}=0,F{row}=""),"",G{row}/(E{row}*{unit_value_cell}))'
    ws[f"H{row}"].number_format = "0%"

    for col_name in COLUMNS:
        ws[f"{COL_LETTERS[col_name]}{row}"].border = BORDER


def _write_section_total(ws, row: int, *, label: str, data_rows: list[int], unit_value_cell: str) -> None:
    ws[f"A{row}"] = label
    ws[f"A{row}"].font = BOLD
    if data_rows:
        first, last = min(data_rows), max(data_rows)
        ws[f"E{row}"] = f"=SUM(E{first}:E{last})"
        ws[f"F{row}"] = f"=SUM(F{first}:F{last})"
        ws[f"G{row}"] = f"=SUM(G{first}:G{last})"
        ws[f"H{row}"] = f'=IF(E{row}=0,"",G{row}/(E{row}*{unit_value_cell}))'
    for col_name in COLUMNS:
        cell = ws[f"{COL_LETTERS[col_name]}{row}"]
        cell.font = BOLD
        cell.fill = SECTION_FILL
        cell.border = BORDER
    ws[f"H{row}"].number_format = "0%"


def build_event_spreadsheet(
    *,
    event_name: str,
    event_date: Optional[str],
    bets: list[dict[str, Any]],
    legs_by_bet_id: dict[int, list[dict[str, Any]]],
    fights: list[tuple[str, str]],
    unit_value: float,
    currency: str,
    output_path: str,
) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Event Recap"

    for col_name in COLUMNS:
        ws.column_dimensions[COL_LETTERS[col_name]].width = 26 if col_name in ("Pick", "Opponent") else 13

    row = 1
    title = event_name + (f", {event_date}" if event_date else "")
    ws[f"A{row}"] = title
    ws[f"A{row}"].font = TITLE_FONT
    row += 2

    # Unit value input cell -- a real input, referenced by every profit
    # formula below, not baked into each row.
    ws[f"A{row}"] = "Unit Value"
    ws[f"A{row}"].font = BOLD
    ws[f"B{row}"] = unit_value
    ws[f"B{row}"].font = INPUT_FONT
    ws[f"B{row}"].number_format = f'"{currency} "#,##0.00'
    ws[f"C{row}"] = f"(Cash amounts below are in {currency})"
    unit_value_cell = f"$B${row}"
    row += 2

    grouped: dict[str, list[dict[str, Any]]] = {"Straight Pick": [], "Prop": [], "Parlay": []}
    for bet in bets:
        legs = legs_by_bet_id.get(bet["id"], [])
        grouped[_leg_category(bet, legs)].append(bet)

    all_section_total_rows: list[int] = []
    section_plural = {"Straight Pick": "Straight Pick", "Prop": "Prop", "Parlay": "Parlay"}

    for section_name, section_bets in grouped.items():
        if not section_bets:
            continue

        ws[f"A{row}"] = section_name + " Bets" if section_name != "Parlay" else "Parlays"
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, size=12)
        row += 1

        _write_header_row(ws, row)
        row += 1

        data_rows: list[int] = []

        for bet in section_bets:
            legs = legs_by_bet_id.get(bet["id"], [])
            status = bet.get("status")
            overall_result = "W" if status == "won" else "L" if status == "loss" else None

            if section_name == "Parlay":
                for leg in legs:
                    fighter = leg.get("fighter_pick")
                    opponent = _find_opponent(fighter, fights) if fighter else ""
                    leg_result = (
                        "W" if leg.get("status") == "won" else "L" if leg.get("status") == "loss" else None
                    )
                    _write_pick_row(
                        ws, row,
                        pick=leg["description"], opponent=opponent, result=leg_result,
                        odds=None, units=None, unit_value_cell=unit_value_cell,
                    )
                    row += 1
                # Combined parlay line -- the bet's own real odds/stake.
                _write_pick_row(
                    ws, row,
                    pick="Parlay Odds", opponent="", result=overall_result,
                    odds=bet.get("odds"), units=bet.get("units"), unit_value_cell=unit_value_cell,
                )
                data_rows.append(row)
                row += 1
            else:
                leg = legs[0] if legs else {"description": bet.get("bet_title") or "Bet", "fighter_pick": None}
                if section_name == "Straight Pick":
                    fighter = leg.get("fighter_pick")
                    pick_text = fighter or leg["description"]
                    opponent = _find_opponent(fighter, fights)
                else:
                    fighter = leg.get("fighter_pick")
                    pick_text = leg["description"]
                    opponent = _find_opponent(fighter, fights) if fighter else ""

                _write_pick_row(
                    ws, row,
                    pick=pick_text, opponent=opponent, result=overall_result,
                    odds=bet.get("odds"), units=bet.get("units"), unit_value_cell=unit_value_cell,
                )
                data_rows.append(row)
                row += 1

        total_label = f"{section_plural[section_name]} Event Totals:"
        _write_section_total(ws, row, label=total_label, data_rows=data_rows, unit_value_cell=unit_value_cell)
        all_section_total_rows.append(row)
        row += 2

    if all_section_total_rows:
        ws[f"A{row}"] = "Event Totals (All Sections):"
        ws[f"A{row}"].font = BOLD
        e_refs = "+".join(f"E{r}" for r in all_section_total_rows)
        f_refs = "+".join(f"F{r}" for r in all_section_total_rows)
        g_refs = "+".join(f"G{r}" for r in all_section_total_rows)
        ws[f"E{row}"] = f"={e_refs}"
        ws[f"F{row}"] = f"={f_refs}"
        ws[f"G{row}"] = f"={g_refs}"
        ws[f"H{row}"] = f'=IF(E{row}=0,"",G{row}/(E{row}*{unit_value_cell}))'
        ws[f"H{row}"].number_format = "0%"
        for col_name in COLUMNS:
            cell = ws[f"{COL_LETTERS[col_name]}{row}"]
            cell.font = BOLD
            cell.fill = TOTAL_FILL
            cell.border = BORDER

    # Assumption note, per the skill's "document every assumption" rule.
    row += 2
    ws[f"A{row}"] = (
        "Note: parlay leg rows show Pick/Opponent/Result only -- individual leg odds "
        "aren't stored, only the combined bet odds shown on the \"Parlay Odds\" row. "
        "A leg's Result is blank unless it was auto-graded (see !eventstart); manually "
        "settling a bet doesn't back-fill its individual legs' results."
    )
    ws[f"A{row}"].font = Font(name=FONT_NAME, italic=True, size=9, color="666666")
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=len(COLUMNS))
    ws[f"A{row}"].alignment = Alignment(wrap_text=True)

    wb.save(output_path)